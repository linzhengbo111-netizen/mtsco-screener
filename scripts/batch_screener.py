#!/usr/bin/env python3
"""
MTSCO 客户背调系统 — 核心逻辑模块
Batch Customer Screener for MTSCO

功能:
  - DuckDuckGo 搜索 + 网页抓取（免费，不需浏览器）
  - DeepSeek API 分析（OpenAI SDK 兼容）
  - Excel 列名自动识别 + 格式校验
  - 分批处理（20家/批）+ 重试 + 去重
  - 成本估算 + 进度回调

用法:
  from batch_screener import BatchScreener
  screener = BatchScreener(api_key="sk-...")
  result = screener.search_company("SK E&C", "South Korea")
"""

import json
import re
import time
import hashlib
from pathlib import Path
from datetime import datetime
from typing import Any, Callable, Optional
from dataclasses import dataclass, field

import requests
import yaml
import pandas as pd
from bs4 import BeautifulSoup
from openai import OpenAI

# ---------------------------------------------------------------------------
# DuckDuckGo 搜索库导入（兼容 v8.x 和 v9+）
# ---------------------------------------------------------------------------
try:
    from ddgs import DDGS  # duckduckgo_search v9+
except ImportError:
    try:
        from duckduckgo_search import DDGS  # duckduckgo_search v8.x
    except ImportError:
        DDGS = None

# ---------------------------------------------------------------------------
# 常量
# ---------------------------------------------------------------------------
SEARCH_TIMEOUT = 30          # DuckDuckGo 搜索超时（秒）
PAGE_FETCH_TIMEOUT = 15      # 网页抓取超时（秒）
MAX_PAGE_LENGTH = 5000       # 单页最大字符数
TOP_N_PAGES = 2              # 抓取前N个网页
MAX_SEARCH_RESULTS = 5       # 每次搜索返回结果数
DELAY_BETWEEN_FETCHES = 1.0  # 抓取间隔（秒），避免被目标站封
DDG_RATE_LIMIT_PAUSE = 30    # DDG限流时暂停秒数
API_RETRY_COUNT = 3          # API调用重试次数
API_RETRY_DELAY = 3          # 重试间隔（秒）
BATCH_SIZE = 20              # 每批处理公司数
API_TIMEOUT = 60             # DeepSeek API 超时（秒）

# 排除的域名（目录站、黄页、B2B平台——信息质量低）
EXCLUDED_DOMAINS = [
    "alibaba.com", "made-in-china.com", "tradeindia.com",
    "exportersindia.com", "ec21.com", "ecplaza.net",
    "indiamart.com", "en.china.cn", "globalsources.com",
    "b2bmap.com", "yellowpages.com", "yelp.com",
    "facebook.com", "instagram.com", "twitter.com", "youtube.com",
    "wikipedia.org", "bloomberg.com", "reuters.com",
    "crunchbase.com", "zoominfo.com", "dnb.com",
    "kompass.com", "europages.com", "werliefernwas.de",
]

# 列名自动识别映射
COMPANY_NAME_ALIASES = [
    "公司名称", "公司名", "company", "company name", "name",
    "客户名称", "客户", "customer", "企业名称", "企业",
    "customer name", "organization", "organisation", "客户公司",
    "companyname", "客户公司名", "company_name",
]

COUNTRY_ALIASES = [
    "国家", "country", "nation", "地区", "region",
    "country/region", "country / region", "国家/地区",
    "国家地区", "nation/region", "location",
]

WEBSITE_ALIASES = [
    "网站", "website", "web", "官网", "网址", "url",
    "site", "主页", "homepage", "web site", "公司网站",
]

# 国家名 → 英文名映射
COUNTRY_MAP = {
    "韩国": "South Korea", "south korea": "South Korea", "korea": "South Korea",
    "kor": "South Korea", "大韩民国": "South Korea", "한국": "South Korea",
    "日本": "Japan", "japan": "Japan", "jpn": "Japan", "日本国": "Japan",
    "中国": "China", "china": "China", "chn": "China", "中国大陆": "China",
    "台湾": "Taiwan", "taiwan": "Taiwan", "twn": "Taiwan",
    "香港": "Hong Kong", "hong kong": "Hong Kong", "hkg": "Hong Kong",
    "新加坡": "Singapore", "singapore": "Singapore", "sgp": "Singapore",
    "马来西亚": "Malaysia", "malaysia": "Malaysia", "mys": "Malaysia",
    "泰国": "Thailand", "thailand": "Thailand", "tha": "Thailand",
    "越南": "Vietnam", "vietnam": "Vietnam", "vnm": "Vietnam",
    "印度尼西亚": "Indonesia", "indonesia": "Indonesia", "idn": "Indonesia",
    "菲律宾": "Philippines", "philippines": "Philippines", "phl": "Philippines",
    "印度": "India", "india": "India", "ind": "India",
    "阿联酋": "UAE", "uae": "UAE", "united arab emirates": "UAE", "are": "UAE",
    "沙特阿拉伯": "Saudi Arabia", "saudi arabia": "Saudi Arabia", "sau": "Saudi Arabia",
    "卡塔尔": "Qatar", "qatar": "Qatar", "qat": "Qatar",
    "科威特": "Kuwait", "kuwait": "Kuwait", "kwt": "Kuwait",
    "阿曼": "Oman", "oman": "Oman", "omn": "Oman",
    "巴林": "Bahrain", "bahrain": "Bahrain", "bhr": "Bahrain",
    "美国": "USA", "usa": "USA", "united states": "USA", "us": "USA",
    "加拿大": "Canada", "canada": "Canada", "can": "Canada",
    "墨西哥": "Mexico", "mexico": "Mexico", "mex": "Mexico",
    "巴西": "Brazil", "brazil": "Brazil", "bra": "Brazil",
    "英国": "UK", "uk": "UK", "united kingdom": "UK", "gbr": "UK",
    "德国": "Germany", "germany": "Germany", "deu": "Germany",
    "法国": "France", "france": "France", "fra": "France",
    "意大利": "Italy", "italy": "Italy", "ita": "Italy",
    "西班牙": "Spain", "spain": "Spain", "esp": "Spain",
    "荷兰": "Netherlands", "netherlands": "Netherlands", "nld": "Netherlands",
    "俄罗斯": "Russia", "russia": "Russia", "rus": "Russia",
    "澳大利亚": "Australia", "australia": "Australia", "aus": "Australia",
    "土耳其": "Turkey", "turkey": "Turkey", "tur": "Turkey",
    "埃及": "Egypt", "egypt": "Egypt", "egy": "Egypt",
    "尼日利亚": "Nigeria", "nigeria": "Nigeria", "nga": "Nigeria",
}

# ---------------------------------------------------------------------------
# 数据类
# ---------------------------------------------------------------------------

@dataclass
class CompanyInput:
    """标准化后的输入公司数据"""
    name: str
    country: str          # 统一为英文名
    website: str = ""
    original_name: str = ""
    original_country: str = ""

@dataclass
class CompanyResult:
    """单家公司分析结果（对应输出Excel的9列）"""
    company_name: str
    country: str
    website: str
    company_type: str = ""
    industry: str = ""
    customs_data_found: str = "no"
    customs_summary: str = ""
    match_level: str = ""
    match_reason: str = ""
    conclusion: str = ""
    confidence: str = ""
    # 额外元数据（不在9列中，供详情查询用）
    search_keywords: list = field(default_factory=list)
    search_results: list = field(default_factory=list)
    page_texts: list = field(default_factory=list)
    raw_ai_response: str = ""
    error_message: str = ""
    analysis_time: str = ""

    def to_dict(self) -> dict:
        return {
            "公司名": self.company_name,
            "国家": self.country,
            "网站": self.website,
            "公司类型": self.company_type,
            "行业": self.industry,
            "海关记录": self.customs_summary if self.customs_summary else "未查到",
            "匹配度": self.match_level,
            "理由": self.match_reason,
            "结论": self.conclusion,
        }

    def to_detail_dict(self) -> dict:
        """详情查询用的完整字典"""
        d = self.to_dict()
        d.update({
            "搜索关键词": ", ".join(self.search_keywords) if self.search_keywords else "",
            "搜索结果摘要": json.dumps(self.search_results, ensure_ascii=False, indent=2) if self.search_results else "",
            "抓取页面": json.dumps(self.page_texts, ensure_ascii=False, indent=2) if self.page_texts else "",
            "AI原始响应": self.raw_ai_response,
            "置信度": self.confidence,
            "错误信息": self.error_message,
            "分析时间": self.analysis_time,
        })
        return d


# ---------------------------------------------------------------------------
# BatchScreener 核心类
# ---------------------------------------------------------------------------

class BatchScreener:
    """批量客户背调器"""

    def __init__(self, api_key: str, model: str = "deepseek-chat"):
        """
        初始化背调器。

        Args:
            api_key: DeepSeek API Key
            model: 模型名 — 'deepseek-chat' (V3) 或 'deepseek-reasoner' (R1)
        """
        if not api_key or not api_key.strip():
            raise ValueError("API Key 不能为空")

        self.api_key = api_key.strip()
        self.model = model
        self._client = None
        self._prompts = None
        self._search_cache = {}  # 搜索缓存，避免同公司重复搜索
        self._stop_flag = False   # 停止标志

    # -----------------------------------------------------------------------
    # DeepSeek Client（懒加载）
    # -----------------------------------------------------------------------
    @property
    def client(self) -> OpenAI:
        if self._client is None:
            self._client = OpenAI(
                api_key=self.api_key,
                base_url="https://api.deepseek.com",
                timeout=API_TIMEOUT,
                max_retries=0,  # 我们自己实现重试
            )
        return self._client

    # -----------------------------------------------------------------------
    # 提示词加载
    # -----------------------------------------------------------------------
    @property
    def prompts(self) -> dict:
        if self._prompts is None:
            config_path = Path(__file__).parent / "config" / "screener_prompt.yaml"
            if not config_path.exists():
                raise FileNotFoundError(f"提示词配置文件不存在: {config_path}")
            with open(config_path, "r", encoding="utf-8") as f:
                self._prompts = yaml.safe_load(f)
        return self._prompts

    # -----------------------------------------------------------------------
    # API Key 验证
    # -----------------------------------------------------------------------
    def validate_api_key(self) -> tuple[bool, str]:
        """
        验证 DeepSeek API Key 是否有效。
        发送一个最小请求测试连通性，不产生实际费用。

        Returns:
            (is_valid, message)
        """
        try:
            # 用最小的请求验证连通性
            resp = self.client.chat.completions.create(
                model=self.model,
                messages=[{"role": "user", "content": "hi"}],
                max_tokens=1,
                temperature=0,
            )
            if resp.choices and len(resp.choices) > 0:
                return True, "API Key 验证通过"
            return False, "API Key 验证失败: 空响应"
        except Exception as e:
            msg = str(e)
            if "401" in msg or "unauthorized" in msg.lower() or "authentication" in msg.lower():
                return False, "API Key 无效 (401 Unauthorized)，请检查 Key 是否正确"
            elif "402" in msg or "insufficient" in msg.lower() or "balance" in msg.lower():
                return False, "账户余额不足 (402)，请充值后重试"
            elif "429" in msg or "rate" in msg.lower():
                return False, "API 请求过于频繁 (429)，请稍后重试"
            elif "timeout" in msg.lower() or "timed out" in msg.lower():
                return False, "API 连接超时，请检查网络"
            else:
                return False, f"API 验证失败: {msg[:200]}"

    # -----------------------------------------------------------------------
    # 搜索
    # -----------------------------------------------------------------------
    def _is_excluded_url(self, url: str) -> bool:
        """检查 URL 是否属于应排除的低质量域名"""
        url_lower = url.lower()
        return any(domain in url_lower for domain in EXCLUDED_DOMAINS)

    def _search_ddg(self, query: str, max_results: int = MAX_SEARCH_RESULTS) -> list[dict]:
        """
        执行 DuckDuckGo 搜索。

        Args:
            query: 搜索关键词
            max_results: 返回结果数

        Returns:
            [{"title": ..., "url": ..., "snippet": ...}, ...]
        """
        if DDGS is None:
            raise ImportError("DuckDuckGo 搜索库未安装。请运行: pip install duckduckgo-search")

        results = []
        try:
            with DDGS() as ddgs:
                raw = list(ddgs.text(query, max_results=max_results))
                for r in raw:
                    url = r.get("href") or r.get("link") or ""
                    if self._is_excluded_url(url):
                        continue
                    results.append({
                        "title": r.get("title", ""),
                        "url": url,
                        "snippet": r.get("body", "") or r.get("snippet", ""),
                    })
            return results
        except Exception as e:
            msg = str(e).lower()
            if "rate" in msg or "429" in msg or "ratelimit" in msg:
                raise RuntimeError(f"DDG_RATELIMIT: {e}")
            if "timeout" in msg or "timed out" in msg:
                raise RuntimeError(f"DDG_TIMEOUT: {e}")
            # 其他错误返回已有结果
            return results

    def _fetch_page(self, url: str) -> str:
        """
        抓取网页文本内容。

        Args:
            url: 网页 URL

        Returns:
            提取的文本内容（截取前 MAX_PAGE_LENGTH 字符）
        """
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/131.0.0.0 Safari/537.36"
            ),
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
        }
        try:
            resp = requests.get(url, headers=headers, timeout=PAGE_FETCH_TIMEOUT, allow_redirects=True)
            if resp.status_code != 200:
                return f"[HTTP {resp.status_code}]"

            # 编码检测
            content_type = resp.headers.get("Content-Type", "")
            if "charset=" in content_type:
                resp.encoding = content_type.split("charset=")[-1].strip()
            else:
                # 尝试自动检测
                resp.encoding = resp.apparent_encoding or "utf-8"

            soup = BeautifulSoup(resp.text, "html.parser")

            # 去除无效标签
            for tag in soup(["script", "style", "nav", "footer", "header", "aside", "noscript", "iframe"]):
                tag.decompose()

            text = soup.get_text(separator="\n", strip=True)
            # 压缩空白行
            text = re.sub(r'\n{3,}', '\n\n', text)

            if len(text) > MAX_PAGE_LENGTH:
                text = text[:MAX_PAGE_LENGTH] + "\n...[truncated]"

            return text if text.strip() else "[页面无文本内容]"

        except requests.exceptions.Timeout:
            return f"[超时: {PAGE_FETCH_TIMEOUT}s]"
        except requests.exceptions.ConnectionError:
            return "[连接失败]"
        except Exception as e:
            return f"[抓取错误: {str(e)[:100]}]"

    def search_company(self, name: str, country: str) -> dict:
        """
        搜索一家公司的信息（3角度 + LinkedIn + 网页抓取）。

        搜索策略:
          1. 公司名 + procurement/supplier
          2. 公司名 + country + industry
          3. site:linkedin.com/company + 公司名

        搜不到则用简称重试。

        Args:
            name: 公司名
            country: 国家（英文）

        Returns:
            {
                "search_results": [...],
                "page_texts": [...],
                "search_keywords": [...],
                "final_query": "...",
            }
        """
        cache_key = hashlib.md5(f"{name}|{country}".encode()).hexdigest()
        if cache_key in self._search_cache:
            return self._search_cache[cache_key]

        all_results = []
        all_texts = []
        keywords_used = []

        # --- 搜索角度 1: 公司名 + procurement ---
        q1 = f'"{name}" procurement OR supplier OR import'
        keywords_used.append(q1)
        try:
            results1 = self._search_ddg(q1)
            all_results.extend(results1)
        except RuntimeError:
            pass  # 限流或超时，继续下一个角度

        # --- 搜索角度 2: 公司名 + country ---
        q2 = f'"{name}" {country} company profile'
        keywords_used.append(q2)
        try:
            results2 = self._search_ddg(q2)
            # 合并并去重
            seen_urls = {r["url"] for r in all_results}
            for r in results2:
                if r["url"] not in seen_urls:
                    all_results.append(r)
                    seen_urls.add(r["url"])
        except RuntimeError:
            pass

        # --- 搜索角度 3: LinkedIn ---
        q3 = f'site:linkedin.com/company "{name}"'
        keywords_used.append(q3)
        try:
            results3 = self._search_ddg(q3)
            seen_urls = {r["url"] for r in all_results}
            for r in results3:
                if r["url"] not in seen_urls:
                    all_results.append(r)
                    seen_urls.add(r["url"])
        except RuntimeError:
            pass

        # --- 如果完全没结果，尝试用简称重搜 ---
        if len(all_results) == 0:
            # 简化公司名：去掉 Co., Ltd., Inc., Corp 等后缀
            short_name = re.sub(
                r'\s+(Co\.?,?\s*)?(Ltd\.?|Limited|Inc\.?|Corp\.?|Corporation|LLC|Pte\.?\s*Ltd\.?|S\.A\.?|GmbH|S\.R\.L\.?).*$',
                '', name, flags=re.IGNORECASE
            ).strip()

            if short_name and len(short_name) >= 4 and short_name != name:
                q4 = f'"{short_name}" {country} company'
                keywords_used.append(f"{q4} [简称重搜]")
                try:
                    results4 = self._search_ddg(q4)
                    all_results.extend(results4)
                except RuntimeError:
                    pass

        # --- 限制结果数 ---
        all_results = all_results[:MAX_SEARCH_RESULTS]

        # --- 抓取前 N 个网页 ---
        for r in all_results[:TOP_N_PAGES]:
            url = r.get("url", "")
            if url:
                try:
                    time.sleep(DELAY_BETWEEN_FETCHES)
                    text = self._fetch_page(url)
                    if text and not text.startswith("[") and text != "[页面无文本内容]":
                        all_texts.append(f"=== {url} ===\n{text}")
                except Exception:
                    continue

        # --- 如果没有搜到有效网页，也尝试抓搜索结果中第一个非排除域名的页面 ---
        if len(all_texts) == 0 and all_results:
            for r in all_results:
                url = r.get("url", "")
                if url and not self._is_excluded_url(url):
                    try:
                        time.sleep(DELAY_BETWEEN_FETCHES)
                        text = self._fetch_page(url)
                        if text and not text.startswith("["):
                            all_texts.append(f"=== {url} ===\n{text}")
                            break
                    except Exception:
                        continue

        result = {
            "search_results": all_results,
            "page_texts": all_texts,
            "search_keywords": keywords_used,
            "final_query": keywords_used[0] if keywords_used else "",
        }

        self._search_cache[cache_key] = result
        return result

    # -----------------------------------------------------------------------
    # DeepSeek 分析
    # -----------------------------------------------------------------------
    def analyze_company(
        self,
        name: str,
        country: str,
        website: str = "",
        search_data: Optional[dict] = None,
    ) -> CompanyResult:
        """
        调用 DeepSeek API 分析一家公司。

        Args:
            name: 公司名
            country: 国家（英文）
            website: 网站URL（选填）
            search_data: search_company() 的返回结果，如果为None则不提供搜索材料

        Returns:
            CompanyResult
        """
        result = CompanyResult(
            company_name=name,
            country=country,
            website=website,
        )

        # 构建搜索材料文本
        search_materials = "未提供搜索材料（仅基于公司名和国家判断）"
        if search_data:
            search_keywords = search_data.get("search_keywords", [])
            search_results = search_data.get("search_results", [])
            page_texts = search_data.get("page_texts", [])

            result.search_keywords = search_keywords
            result.search_results = search_results
            result.page_texts = page_texts

            parts = []
            if search_results:
                parts.append("=== DuckDuckGo 搜索结果 ===")
                for i, sr in enumerate(search_results, 1):
                    parts.append(f"{i}. {sr['title']}\n   URL: {sr['url']}\n   摘要: {sr['snippet']}")

            if page_texts:
                parts.append("\n=== 抓取的网页内容 ===")
                for pt in page_texts:
                    parts.append(pt[:3000])  # 每页最多取3000字符

            search_materials = "\n\n".join(parts) if parts else "搜索无结果"

        # 构建完整提示
        system_prompt = self.prompts.get("system_prompt", "")
        user_template = self.prompts.get("user_prompt_template", "")

        user_message = user_template.format(
            company_name=name,
            country=country,
            website=website or "未提供",
            search_materials=search_materials[:12000],  # 限制总长度
        )

        # 调用 DeepSeek API（含重试）
        raw_json = None
        last_error = ""

        for attempt in range(1, API_RETRY_COUNT + 1):
            if self._stop_flag:
                result.error_message = "用户手动停止"
                return result

            try:
                response = self.client.chat.completions.create(
                    model=self.model,
                    messages=[
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": user_message},
                    ],
                    temperature=0.1,  # 低温度 = 更一致
                    max_tokens=2000,
                    response_format={"type": "json_object"},
                )

                raw_text = response.choices[0].message.content.strip()
                result.raw_ai_response = raw_text

                # 解析 JSON
                # 处理可能的 markdown 包裹
                json_text = raw_text
                if json_text.startswith("```"):
                    json_text = re.sub(r'^```(?:json)?\s*', '', json_text)
                    json_text = re.sub(r'\s*```$', '', json_text)

                raw_json = json.loads(json_text)
                break  # 成功，跳出重试循环

            except json.JSONDecodeError as e:
                last_error = f"JSON解析失败(第{attempt}次): {e}"
                if attempt < API_RETRY_COUNT:
                    time.sleep(API_RETRY_DELAY)
            except Exception as e:
                last_error = str(e)
                msg_lower = last_error.lower()

                # 余额不足 —— 不重试，直接抛出让上层处理
                if "402" in msg_lower or "insufficient" in msg_lower or "balance" in msg_lower:
                    raise RuntimeError(f"BALANCE_INSUFFICIENT: {last_error}")

                # 超时 —— 按计划重试
                if "timeout" in msg_lower or "timed out" in msg_lower:
                    if attempt < API_RETRY_COUNT:
                        time.sleep(API_RETRY_DELAY * attempt)
                        continue

                # 其他错误
                if attempt < API_RETRY_COUNT:
                    time.sleep(API_RETRY_DELAY)
                else:
                    last_error = f"API调用失败({API_RETRY_COUNT}次重试后): {last_error}"

        # --- 解析结果 ---
        if raw_json:
            result.company_type = str(raw_json.get("company_type", "uncertain"))
            result.industry = str(raw_json.get("industry", "unknown"))
            result.customs_data_found = str(raw_json.get("customs_data_found", "no"))
            result.customs_summary = str(raw_json.get("customs_summary", ""))
            result.match_level = str(raw_json.get("match_level", "medium"))
            result.match_reason = str(raw_json.get("match_reason", ""))
            result.conclusion = str(raw_json.get("conclusion", "需人工核实"))
            result.confidence = str(raw_json.get("confidence", "medium"))
            result.website = str(raw_json.get("website", "")) or website or ""
            # search_summary 存到 error_message 用于调试（不影响9列输出）
            if raw_json.get("search_summary"):
                result.raw_ai_response += f"\n\n--- SEARCH SUMMARY ---\n{raw_json['search_summary']}"
        else:
            result.company_type = "uncertain"
            result.industry = "unknown"
            result.match_level = "medium"
            result.match_reason = f"AI分析失败: {last_error}"
            result.conclusion = "需人工核实"
            result.confidence = "low"
            result.error_message = last_error

        result.analysis_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        return result

    # -----------------------------------------------------------------------
    # Excel 校验
    # -----------------------------------------------------------------------
    def validate_excel(self, df: pd.DataFrame) -> tuple:
        """
        校验并标准化上传的 Excel。

        1. 列名自动识别
        2. 检查必填列是否有值
        3. 国家名统一转英文

        Args:
            df: pandas DataFrame

        Returns:
            (is_valid, error_messages, standardized_df_or_None)
        """
        errors = []

        if df.empty:
            return False, ["Excel 文件为空，没有数据行"], None

        # --- 列名识别 ---
        col_map = {}
        cols_lower = {c.lower().strip(): c for c in df.columns}

        # 找公司名列
        for alias in COMPANY_NAME_ALIASES:
            if alias in cols_lower:
                col_map["name"] = cols_lower[alias]
                break
        if "name" not in col_map:
            return False, [
                f"未识别到「公司名」列。请确保 Excel 包含以下列名之一: {', '.join(COMPANY_NAME_ALIASES[:6])}"
            ], None

        # 找国家列
        for alias in COUNTRY_ALIASES:
            if alias in cols_lower:
                col_map["country"] = cols_lower[alias]
                break
        if "country" not in col_map:
            return False, [
                f"未识别到「国家」列。请确保 Excel 包含以下列名之一: {', '.join(COUNTRY_ALIASES[:5])}"
            ], None

        # 找网站列（可选）
        for alias in WEBSITE_ALIASES:
            if alias in cols_lower:
                col_map["website"] = cols_lower[alias]
                break

        # --- 数据提取 ---
        names = df[col_map["name"]].astype(str).str.strip()
        countries = df[col_map["country"]].astype(str).str.strip()
        websites = df[col_map.get("website", "")].astype(str).str.strip() if "website" in col_map else pd.Series([""] * len(df))

        # --- 逐行校验 ---
        valid_rows = []
        for i, (n, c, w) in enumerate(zip(names, countries, websites)):
            row_num = i + 2  # Excel 行号（第1行是表头）
            row_errors = []

            if not n or n.lower() in ("nan", "none", "null", ""):
                row_errors.append(f"第{row_num}行: 公司名为空")
            if not c or c.lower() in ("nan", "none", "null", ""):
                row_errors.append(f"第{row_num}行: 国家为空")

            if row_errors:
                errors.extend(row_errors)
                continue

            # 统一国家名
            c_lower = c.lower().strip()
            c_english = COUNTRY_MAP.get(c_lower, c)

            valid_rows.append({
                "name": n,
                "original_name": n,
                "country": c_english,
                "original_country": c,
                "website": w if w and w.lower() not in ("nan", "none", "null") else "",
            })

        if errors:
            return False, errors, None

        if not valid_rows:
            return False, ["未检测到有效数据行，请检查 Excel 内容"], None

        result_df = pd.DataFrame(valid_rows)
        return True, [], result_df

    # -----------------------------------------------------------------------
    # 去重
    # -----------------------------------------------------------------------
    @staticmethod
    def deduplicate(df: pd.DataFrame, history_df: Optional[pd.DataFrame]) -> tuple:
        """
        严格名称匹配去重。

        Args:
            df: 待处理的公司列表
            history_df: 历史结果 Excel

        Returns:
            (去重后的df, 被去重的公司名列表)
        """
        if history_df is None or history_df.empty:
            return df, []

        # 找历史数据中的公司名列
        history_names = None
        for col in history_df.columns:
            col_lower = col.lower().strip()
            for alias in COMPANY_NAME_ALIASES:
                if col_lower == alias.lower():
                    history_names = set(history_df[col].astype(str).str.strip().str.lower())
                    break
            if history_names:
                break

        if history_names is None:
            # 尝试第一列
            history_names = set(history_df.iloc[:, 0].astype(str).str.strip().str.lower())

        # 严格匹配
        current_names = df["name"].str.lower().str.strip()
        duplicate_mask = current_names.isin(history_names)
        duplicates = df[duplicate_mask]["name"].tolist()
        deduped = df[~duplicate_mask].copy()

        return deduped, duplicates

    # -----------------------------------------------------------------------
    # 成本估算
    # -----------------------------------------------------------------------
    @staticmethod
    def estimate_cost(count: int) -> dict:
        """
        估算分析成本和时间。

        Args:
            count: 公司数量

        Returns:
            {"companies": int, "batches": int, "estimated_minutes": float, "estimated_cost_rmb": float}
        """
        batches = (count + BATCH_SIZE - 1) // BATCH_SIZE
        # 每家公司约3次搜索 + 2次页面抓取 + 1次API调用, ~15-20秒
        estimated_seconds = count * 18
        estimated_minutes = max(1, round(estimated_seconds / 60, 1))

        # DeepSeek V3: ~1元/百万token; 每家约3000-5000 token输入+~500 token输出
        # 成本约0.004-0.006元/家
        cost_per_company = 0.005 if count <= 100 else 0.004
        estimated_cost = round(count * cost_per_company, 2)

        return {
            "companies": count,
            "batches": batches,
            "estimated_minutes": estimated_minutes,
            "estimated_cost_rmb": estimated_cost,
        }

    # -----------------------------------------------------------------------
    # 批量处理
    # -----------------------------------------------------------------------
    def run_batch(
        self,
        companies: list[dict],
        progress_callback: Optional[Callable[[str, dict], None]] = None,
    ) -> list[CompanyResult]:
        """
        批量处理公司列表（20家/批）。

        流程:
          1. 逐家搜索
          2. 逐家AI分析
          3. 重试失败项
          4. 限流检测 + 自动暂停

        Args:
            companies: 标准化后的公司列表 [{"name":..., "country":..., "website":...}, ...]
            progress_callback: 进度回调，签名 callback(status_type, data)
                status_type: "search_start" | "search_done" | "analysis_start" | "analysis_done" |
                            "company_done" | "batch_done" | "error" | "rate_limit" | "stopped"

        Returns:
            [CompanyResult, ...]
        """
        all_results: list[CompanyResult] = []
        total = len(companies)
        self._stop_flag = False

        for batch_idx in range(0, total, BATCH_SIZE):
            if self._stop_flag:
                if progress_callback:
                    progress_callback("stopped", {"completed": len(all_results), "total": total})
                break

            batch = companies[batch_idx:batch_idx + BATCH_SIZE]
            batch_num = batch_idx // BATCH_SIZE + 1
            total_batches = (total + BATCH_SIZE - 1) // BATCH_SIZE

            for i, company in enumerate(batch):
                if self._stop_flag:
                    if progress_callback:
                        progress_callback("stopped", {
                            "completed": len(all_results), "total": total
                        })
                    break

                global_idx = batch_idx + i + 1
                name = company["name"]
                country = company["country"]
                website = company.get("website", "")

                # --- 步骤1: 搜索 ---
                if progress_callback:
                    progress_callback("search_start", {
                        "company": name, "index": global_idx, "total": total,
                        "batch": batch_num, "total_batches": total_batches,
                    })

                search_data = None
                try:
                    search_data = self.search_company(name, country)
                except RuntimeError as e:
                    msg = str(e)
                    if "DDG_RATELIMIT" in msg:
                        if progress_callback:
                            progress_callback("rate_limit", {"message": "DuckDuckGo 限流，暂停30秒..."})
                        time.sleep(DDG_RATE_LIMIT_PAUSE)
                        try:
                            search_data = self.search_company(name, country)
                        except RuntimeError:
                            pass  # 仍失败，用无搜索数据分析
                    # 其他搜索错误：继续用无搜索数据分析

                if progress_callback:
                    progress_callback("search_done", {
                        "company": name,
                        "results_found": len(search_data.get("search_results", [])) if search_data else 0,
                        "pages_fetched": len(search_data.get("page_texts", [])) if search_data else 0,
                    })

                # --- 步骤2: AI 分析 ---
                if progress_callback:
                    progress_callback("analysis_start", {
                        "company": name, "index": global_idx, "total": total,
                    })

                try:
                    result = self.analyze_company(name, country, website, search_data)
                except RuntimeError as e:
                    msg = str(e)
                    if "BALANCE_INSUFFICIENT" in msg:
                        result = CompanyResult(
                            company_name=name, country=country, website=website,
                            error_message="API余额不足，请充值后重试",
                            match_level="medium", conclusion="需人工核实",
                            match_reason="分析中断: API余额不足",
                            confidence="low",
                        )
                        if progress_callback:
                            progress_callback("error", {
                                "company": name,
                                "error": "API余额不足，请充值后点击'重试失败项'继续",
                                "type": "balance",
                            })
                        # 余额不足：不继续了，返回已有结果
                        self._stop_flag = True
                        all_results.append(result)
                        break
                    else:
                        result = CompanyResult(
                            company_name=name, country=country, website=website,
                            error_message=f"API调用失败: {msg[:200]}",
                            match_level="medium", conclusion="需人工核实",
                            match_reason="分析失败: API调用异常",
                            confidence="low",
                        )

                result.analysis_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                all_results.append(result)

                if progress_callback:
                    progress_callback("company_done", {
                        "result": result,
                        "index": global_idx, "total": total,
                    })

            if self._stop_flag:
                break

            if progress_callback:
                progress_callback("batch_done", {
                    "batch": batch_num, "total_batches": total_batches,
                    "completed": len(all_results), "total": total,
                })

        return all_results

    def stop(self):
        """停止批量分析"""
        self._stop_flag = True

    # -----------------------------------------------------------------------
    # 详情查询
    # -----------------------------------------------------------------------
    @staticmethod
    def get_company_detail(results: list, company_name: str) -> Optional[dict]:
        """
        根据公司名查找详情（模糊匹配）。

        Args:
            results: 分析结果列表
            company_name: 要查询的公司名

        Returns:
            详情字典或None
        """
        name_lower = company_name.lower().strip()
        for r in results:
            if name_lower in r.company_name.lower():
                return r.to_detail_dict()
        return None


# ---------------------------------------------------------------------------
# 工具函数: Excel 导出
# ---------------------------------------------------------------------------

def results_to_excel(results: list[CompanyResult], output_path: str) -> str:
    """
    将分析结果导出为带颜色标记的 Excel 文件。

    Args:
        results: 分析结果列表
        output_path: 输出文件路径

    Returns:
        输出文件路径
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    # 按匹配度排序: high → medium → low
    match_order = {"high": 0, "medium": 1, "low": 2}
    sorted_results = sorted(results, key=lambda r: match_order.get(r.match_level, 99))

    # 创建 workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "背调结果"

    # 表头
    headers = ["公司名", "国家", "网站", "公司类型", "行业", "海关记录", "匹配度", "理由", "结论"]
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # 颜色定义
    green_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    yellow_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    body_font = Font(name="Arial", size=10)
    body_alignment = Alignment(vertical="top", wrap_text=True)

    # 写入数据
    for row_idx, r in enumerate(sorted_results, 2):
        d = r.to_dict()
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=d.get(header, ""))
            cell.font = body_font
            cell.alignment = body_alignment
            cell.border = thin_border

            # 匹配度列颜色
            if header == "匹配度":
                cell.alignment = Alignment(horizontal="center", vertical="top")
                match_str = str(d.get("匹配度", "")).lower()
                if match_str == "high":
                    cell.fill = green_fill
                    cell.value = "🟢 高匹配"
                elif match_str == "medium":
                    cell.fill = yellow_fill
                    cell.value = "🟡 中等"
                elif match_str == "low":
                    cell.fill = red_fill
                    cell.value = "🔴 不匹配"

            # 结论列颜色
            if header == "结论":
                conclusion = str(d.get("结论", ""))
                if "发开发信" == conclusion:
                    cell.fill = green_fill
                elif "需人工核实" in conclusion:
                    cell.fill = yellow_fill
                elif "不发" in conclusion:
                    cell.fill = red_fill

    # 列宽
    col_widths = [30, 12, 28, 14, 16, 35, 14, 50, 16]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # 冻结首行
    ws.freeze_panes = "A2"

    # 自动筛选
    ws.auto_filter.ref = f"A1:I{len(sorted_results) + 1}"

    wb.save(output_path)
    return output_path


def create_sample_excel(output_path: str) -> str:
    """创建示例 Excel 文件"""
    sample_data = {
        "公司名称": [
            "SK Engineering & Construction Co., Ltd.",
            "Hyundai Engineering & Construction Co., Ltd.",
            "SAMSUNG HEAVY INDUSTRIES CO., LTD.",
        ],
        "国家": ["South Korea", "South Korea", "South Korea"],
        "网站": ["www.skec.com", "www.hdec.kr", "www.samsungshi.com"],
    }
    df = pd.DataFrame(sample_data)
    df.to_excel(output_path, index=False)
    return output_path
